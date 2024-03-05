from openpyxl import Workbook, load_workbook

def append_to_excel(filename, data):
    """
    Appends data to an Excel file. 
    :param filename: Name of the Excel file
    :param data: List of tuples, where each tuple represents a row of data
    """

    # Try to load the existing workbook
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        
    except FileNotFoundError:
        # If the workbook doesn't exist, create a new one
        workbook = Workbook()
        sheet = workbook.active

    # Append data to the sheet
    for row in data:
        sheet.append(row)

    # Save changes
    workbook.save(filename)

# Sample usage:
data_to_append = [
    ("Timestamp", "Temperature"),  # Example header row
    ("2023-10-15 07:27:07", 25),
    ("2023-10-15 07:30:00", 26),
]
append_to_excel("log_data.xlsx", data_to_append)
