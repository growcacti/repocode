import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import os
import csv
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import openpyxl
from openpyxl import Workbook, load_workbook

class ScrolledText(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent)

        # Create the scrollbar
        self.scrollbar = tk.Scrollbar(self)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Create the Text widget
        self.text_widget = tk.Text(
            self, wrap=tk.WORD, yscrollcommand=self.scrollbar.set, *args, **kwargs
        )
        self.text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Configure the scrollbar to scroll the Text widget
        self.scrollbar.config(command=self.text_widget.yview)
    def insert(self, index, chars, *args):
        self.text_widget.insert(index, chars, *args)

    def delete(self, first, last=None):
        self.text_widget.delete(first, last)

 
 
class LogBook:
    def __init__(self, parent):
        self.parent = parent
        self.nb = ttk.Notebook(parent)
        self.f1 = ttk.Frame(parent)

        self.f2 = ttk.Frame(parent)
        self.f3 = ttk.Frame(parent)
        self.f4 = ttk.Frame(parent)
        self.nb.grid(row=0, column=0)
        self.logfile =  ScrolledText(self.f2)
        self.logfile.grid(row=1, column=1)
        self.view =  ScrolledText(self.f4)
        self.view.grid(row=1, column=1)
               # Automatically load log file if it exist

        self.plot_frame = ttk.Frame(self.f3)
        self.plot_frame.grid(row=1, column=0, padx=10, pady=10)

        self.nb.add(self.f1, text="User Entry")
        self.nb.add(self.f2, text="Logs")
##        self.nb.add(self.f3, text="Temp vs Time")
        self.nb.add(self.f4, text="View")
        self.log_filename = ""
        self.log_entry_count = 0
        self.rx_etm = []
        self.tx_etm = []
        self.comment = ""
        self.event_list = [
            "INITIAL_SPEC",
            "PRE_TV_FUNCT",
            "TV1_HOT_SPEC",
            "TV1_COLD_SPEC",
            "TV4_HOT_SPEC",
            "TV4_HOT_SPEC",
            "POST_TV_FUNCT",
            "PRE_TC_FUNCT",
            "TC1_HOT_SPEC",
            "TC1_COLD_SPEC",
            "TC2_HOT",
            "TC2_COLD",
            "TC3 HOT",
            "TC3 COLD",
            "TC4 HOT",
            "TC4 COLD",
            "TC5 HOT",
            "TC5 COLD",
            "TC6 HOT_FUNCT",
            "TC6 COLD_FUNCT",
            "TC7 HOT",
            "TC7 COLD",
            "TC8 HOT",
            "TC8 COLD",
            "TC9 HOT",
            "TC9 COLD",
            "TC10 HOT_SPEC",
            "TC10 COLD_SPEC",
            "FINAL_SPEC"
            "Transistion to Hot ",
            "Transistion to cold",
            "Transition to bakeout",
            "PRE_VIB_FUNCT",
            "VIB_X_SIN1",
            "VIB_X_RND",
            "VIB_X_SIN2",
            "VIB_Y_SIN1",
            "VIB_Y_RND",
            "VIB_Y_SIN2",
            "VIB_Z_SIN1",
            "VIB_Z_RND",
            "VIB_Z_SIN2",
            "POST_VIB_FUNCT",
            "STACK ALIGNMENT",
            "NONE",
            " ",
        ]

        self.test_list = [
            "6.0.2 Software Load",
            "6.5 Input Power",
            "6.6 voltage Telemetry",
            "6.7 RX Freq Range",
            "6.8 Frequency Range",
            "6.9 TX Frequency",
            "6.10 TX Output Power",
            "6.11 TX Spurious",
            "6.12 Phase Noise",
            "6.13 AM residual ",
            "6.14 TX Carrier Supression",
            "6.18 Modes of operation",
            "6.19 ACQ HOP",
            "6.20 ACQ HOP 2-5",
            "6.21 FSK Thesholds",
            "6.22 Bit Error",
            "6.25 AGC Curve",
            "6.28 TLM CUrve",
            "6.29 1553 Command & Telemerty",
            "6.36 A/B TEST",
            "TX Power Out Alignment",
            "Confidence Test",
            "6.3 Input VSWR",
            "STE CAL",
            "6.15 AM PM",
            "6.16 TX Linearity",
            "6.26 MDU Interface",
            "6.4O RX Monitor",
            "6.41 TX Monitor",
            "6.37 TX Power TLM",
            "6.38 Voltage Thesholds",
            "NONE",
            "Speical Tests",
            "Troubleshooting",
            "In Error",
        ]

        tk.Label(self.f1, text="Serial Number:").grid(row=1, column=0, padx=10, pady=10)
        self.serial_entry = tk.Entry(self.f1)
        self.serial_entry.grid(row=1, column=1, padx=10, pady=10)

        # Buttons
        self.save_button = ttk.Button(
            self.f1, text="Save Serial", command=self.save_serial
        )
        self.save_button.grid(row=2, column=0, padx=10, pady=10)
       
        self.load_button = ttk.Button(
            self.f1, text="Load Serial", command=self.load_serial
        )
        self.load_button.grid(row=2, column=3, padx=10, pady=10)

        self.timestamp = tk.Entry(self.f1)
        self.timestamp.grid(row=0, column=10)
        self.timestamp.insert(tk.END, datetime.now().strftime("%m/%d/%Y %H:%M"))
        self.c1 = ttk.Combobox(self.f1, values=self.event_list)
        self.c1.grid(row=7, column=1)

        self.c2 = ttk.Combobox(self.f1, values=self.test_list)
        self.c2.grid(row=9, column=1)
        tk.Label(self.f1, text="Start Time").grid(row=4, column=3)
        tk.Label(self.f1, text="Stop Time").grid(row=6, column=3)
        tk.Label(self.f1, text="Test Environment").grid(row=7, column=0)
        tk.Label(self.f1, text="Test Number").grid(row=9, column=0)
        self.timestart = ttk.Combobox(self.f1, values=self.time_values())
        self.timestart.grid(row=4, column=4)
        self.timestop = ttk.Combobox(self.f1, values=self.time_values())
        self.timestop.grid(row=6, column=4)
        self.badge = tk.Entry(self.f1)
        self.badge.grid(row=6, column=1)
        self.label1 = tk.Label(self.f1, text="Badge").grid(row=6, column=0)

        tk.Label(self.f1, text="RX ETM").grid(row=22, column=0)
        tk.Label(self.f1, text="TX ETM").grid(row=23, column=0)
        self.rx_etm = tk.Entry(self.f1)
        self.tx_etm = tk.Entry(self.f1)
        self.rx_etm.grid(row=22, column=1)
        self.tx_etm.grid(row=23, column=1)
        tk.Label(self.f1, text="Temperature").grid(row=12, column=0)
        tk.Label(self.f1, text="=--").grid(row=13, column=0)
        tk.Label(self.f1, text="--").grid(row=14, column=0)

        self.temp = ttk.Combobox(self.f1, values=["-24", "25", "61"])
        self.temp.grid(row=12, column=1)

        self.add_entry = tk.Text(self.f1, width=30, height=2)
        self.add_entry.grid(row=14, column=1)
        btn = ttk.Button(self.f1, text="Update Log", command=self.update_log)
        btn.grid(row=29, column=1, padx=10, pady=10)
        btn = ttk.Button(self.f1, text="Clear Entries", command=self.clear_entries)
        btn.grid(row=30, column=3, padx=10, pady=10)
        self.export_button = ttk.Button(
            self.f1, text="----",
        )
        self.export_button.grid(
            row=31, column=3, padx=10, pady=10
        )  # Adjust the grid positioning accordingly

        btn1 = ttk.Button(self.f1, text="grab start time as Now", command=self.start)
        btn1.grid(row=5, column=4, padx=10, pady=10)
        btn2 = ttk.Button(self.f1, text="use stop time as now", command=self.stop)
        btn2.grid(row=7, column=4, padx=10, pady=10)
        self.log_message = "-"
        btn4 = ttk.Button(
            self.f2, text="Load Previous Logs", command=self.load_log_file
        )
        btn4.grid(row=32, column=4, padx=5, pady=5)
        btn44 = ttk.Button(
            self.f2, text="Clear", command=self.clearlog
        )
        btn44.grid(row=34, column=4, padx=5, pady=5) 
        btn5 = ttk.Button(self.f1, text="Plot Time vs Temp")
        btn5.grid(row=33, column=5, padx=5, pady=5)
        self.btnone = ttk.Button(
            self.f4, text="Load Previous View", command=self.load_view_file)
        self.btnone.grid(row=25, column=4, padx=10, pady=10)
        self.btntwo = ttk.Button(
            self.f4, text="Clear", command=self.clearview)
        self.btntwo.grid(row=26, column=4, padx=10, pady=10)
  
    def save_serial(self):
        # Save the serial number to a file
        with open("serial_number.txt", "w") as file:
            file.write(self.serial_entry.get())

    def load_serial(self):
        # Load the serial number from a file
        try:
            with open("serial_number.txt", "r") as file:
                serial = file.read()
                self.serial_entry.delete(0, tk.END)  # Clear the entry
                self.serial_entry.insert(0, serial)  # Insert the loaded serial number
        except FileNotFoundError:
            messagebox.showerror(
                "Error",
                "Serial number file not found. Please enter and save a serial number.",
            )

    def start(self):
        self.timestart.set(datetime.now().strftime("%I:%M %p"))
        return

    def stop(self):
        self.timestop.set(datetime.now().strftime("%I:%M %p"))
        return

    def time_values(self):
        times = []
        postfix = "AM"
        for minute in range(0, 1440):
            if minute >= 720:
                postfix = "PM"
            hour = minute // 60
            minutes = minute % 60
            times.append(f"{hour:02}:{minutes:02} {postfix}")

        return times

    def save_newlog(self):
        self.log_filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("Comma Separated Values", "*.csv"), ("All files", "*.*")],
        )

    def save_newview(self):    
        self.view_filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        )

    def update_log(self):
        self.uut = self.serial_entry.get()
        self.log_entry_count += 1
        
        self.view_message = f"Log Entry Number: {self.log_entry_count}\n"
        self.view_message += "Serial Number:  " + self.serial_entry.get() + "\n"
        self.view_message += "Badge:  " + self.badge.get() + "\n"
        self.view_message += (
            "Timestamp:  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        )
        self.view_message += "Test Environment: " + self.c1.get() + "\n"
        self.view_message += "Test Number: " + self.c2.get() + "\n"
        self.view_message += "Start Time: " + self.timestart.get() + "\n"
        self.view_message += "Stop Time: " + self.timestop.get() + "\n"
        self.view_message += "Temp (C) : " + self.temp.get() + "\n"
        self.view_message += "RX ETM:  " + self.rx_etm.get() + "\n"
        self.view_message += "TX ETM:  " + self.tx_etm.get() + "\n"
        self.view_message += "Comments: " + self.add_entry.get("1.0", tk.END)
        self.view_message += "-" * 40 + "\n"


        self.log_message = f"{self.log_entry_count}" + ","
        self.log_message += (datetime.now().strftime("%Y-%m-%d %H:%M") + ",")
        self.log_message += self.badge.get() + ","
        self.log_message += self.temp.get() + ","
        self.log_message += self.c1.get() + ","
        self.log_message += self.c2.get() + ","
       
       
        self.log_message += self.rx_etm.get() + ","
        self.log_message += self.tx_etm.get() + ","
        self.log_message += self.timestart.get() + ","
        self.log_message +=  self.timestop.get() + "," + "\n"
        self.timestamp =  (datetime.now().strftime("%d-%m-%Y  %H:%M")+ ",")
        self.comment += self.timestamp + self.add_entry.get("1.0", tk.END) + "\n"
    

        # Insert the log message into Tab2's text widget
        self.logfile.insert(tk.END, self.log_message)
        self.view.insert(tk.END, self.view_message)                    
        serial_number = self.serial_entry.get().strip()
        if not serial_number:
            # Optionally, handle the case where the serial number is not provided
            messagebox.showerror("Error", "Please enter a serial number.")
            return
        self.header =(["Entry","Badge","Timestamp","Temperature","Test Environment","Test","RX ETM","TX ETM","Start","Stop"]) 
        self.log_filename = f"log_{serial_number}.csv"
        self.view_filename = f"view_{serial_number}.txt"
               
        # Check if CSV exists, if not write the header
        if not os.path.exists(self.log_filename):
            with open(self.log_filename, 'w', newline='') as file:
                csv_writer = csv.writer(file)
                csv_writer.writerow(self.header)
        if not os.path.exists(self.view_filename):
              with open(self.view_filename, "a") as file:
                  file.write(self.view_message)

           
                  
        # Append the log message to CSV
       
        # Load or create Excel workbook
        excel_file_name = self.log_filename.replace(".csv", ".xlsx")
        if not os.path.exists(excel_file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Logs"
            # Add headers to Excel only if it's a new workbook
            ws.append(self.header)
        else:
            wb = load_workbook(excel_file_name)
            ws = wb["Logs"]
        
        # Directly append new data to Excel, no need to read from CSV
        new_data = self.log_message.strip().split(',')
        ws.append(new_data)

        # Save changes to the Excel workbook
        wb.save(excel_file_name)

    

    def clear_entries(self):
        # Clear the content of the entry widgets
        self.badge.delete(0, tk.END)

       
        self.c1.set("")
        self.c2.set("")
        self.timestart.set("")
        self.timestop.set("")

        self.rx_etm.delete(0, tk.END)
        self.tx_etm.delete(0, tk.END)
        self.temp.set("")

        self.add_entry.delete("1.0", tk.END)

    def load_log_file(self):
        self.log_filename = filedialog.askopenfilename(
            defaultextension=".csv",
            filetypes=[("Comma Separated Values files", "*.csv"), ("All files", "*.*")],
        )
        if not self.log_filename:
            self.save_newlog()
        
        try:
            with open(self.log_filename, "r") as file:
                content = file.read()
                 
                self.logfile.insert(
                    tk.END, content
                )  # Insert the content from the file
                return self.log_filename
           
        except FileNotFoundError:
            messagebox.showerror(
                "Error", "file not found. Please enter and save a log file next window."
            )
            self.save_newlog()
    def load_view_file(self):
        self.view_filename = filedialog.askopenfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        )
        if not self.view_filename:
            self.save_newview()
        
        try:
            with open(self.view_filename, "r") as file:
                content = file.read()
                
                self.view.insert(
                    tk.END, content
                )  # Insert the content from the file
                return self.view_filename
           
        except FileNotFoundError:
            messagebox.showerror(
                "Error", "file not found. Please enter and save a log file next window."
            )
            self.save_newlog()
    def clearlog(self):
        self.logfile.delete("1.0", tk.END)
    def clearview(self):
        self.view.delete("1.0", tk.END)    






    def draw_plot(self):
        timelog = filedialog.askopenfilename(
            defaultextension=".log",
            filetypes=[("Log files", "*.log"), ("All files", "*.*")],
        )

        with open(timelog, "r") as time_file:
            lines = time_file.readlines()
            self.timelist = [
                line.strip() for line in lines if line.strip()
            ]  # Take non-empty lines and remove any whitespace

        templog = filedialog.askopenfilename(
            defaultextension=".log",
            filetypes=[("Log files", "*.log"), ("All files", "*.*")],
        )
        with open(templog, "r") as temp_file:
            lines = temp_file.readlines()
            self.templist = [
                float(line.strip()) for line in lines if line.strip()
            ]  # Convert each non-empty line to a float

        # Ensure that the number of timestamps and temperatures match
        if len(self.timelist) != len(self.templist):
            print(
                "Error: The number of timestamps does not match the number of temperature readings."
            )
            return

        # Convert time_points to a format suitable for plotting
        time_points = np.arange(len(self.timelist))

        fig, ax = plt.subplots(figsize=(10, 8))
        ax.plot(time_points, self.templist, "-o", label="Temperature (°C)")
        ax.set_title("Temperature Progression Over Time")
        ax.set_xlabel("Entries")
        ax.set_ylabel("Temperature (°C)")
        ax.set_xticks(time_points)
        ax.set_xticklabels(
            self.timelist, rotation=90, fontsize=8, ha="right"
        )  # Rotating for better visualization
        ax.legend()
        ax.grid(True)

        canvas = FigureCanvasTkAgg(fig, master=self.plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=1)


if __name__ == "__main__":
    parent = tk.Tk()
    parent.title("Test Log")
    log = LogBook(parent)
    parent.mainloop()
